"""Document parsing orchestrator — routes file types to appropriate parsers."""

import json
import logging
import time
from pathlib import Path

import fitz  # PyMuPDF
from zhipuai import ZhipuAI

from ..config import settings

logger = logging.getLogger(__name__)


# ── Zhipu File Extract API ────────────────────────────────────────────

class ZhipuFileParser:
    def __init__(self):
        self.client = ZhipuAI(api_key=settings.zhipu_api_key, base_url="https://open.bigmodel.cn/api/paas/v4")

    def parse(self, file_path: Path) -> tuple[str, str]:
        try:
            file_object = self.client.files.create(file=file_path, purpose="file-extract")
            content = self._poll_content(file_object.id)
            try:
                self.client.files.delete(file_id=file_object.id)
            except Exception:
                pass
            return content, "success"
        except Exception as e:
            logger.warning(f"Zhipu parse failed for {file_path.name}: {e}")
            return "", f"error: {e}"

    def _poll_content(self, file_id: str, max_retries: int = 60, interval: float = 2.0) -> str:
        for _ in range(max_retries):
            try:
                raw = self.client.files.content(file_id=file_id)
                text = raw.content.decode("utf-8")
                if not text.strip():
                    time.sleep(interval)
                    continue
                try:
                    data = json.loads(text)
                    if isinstance(data, dict) and "content" in data:
                        inner = data["content"]
                        if inner and inner.strip():
                            return inner
                        return "（此文件仅含图片，无文字内容。）"
                except (json.JSONDecodeError, TypeError):
                    pass
                return text
            except Exception:
                pass
            time.sleep(interval)
        raise TimeoutError(f"文件解析超时 (file_id={file_id})")


# ── Excel → Markdown ──────────────────────────────────────────────────

class XlsParser:
    def parse(self, file_path: Path) -> tuple[str, str]:
        ext = file_path.suffix.lower()
        try:
            if ext == ".xls":
                return self._parse_xls(file_path)
            else:
                return self._parse_xlsx(file_path)
        except Exception as e:
            logger.warning(f"Excel parse failed for {file_path.name}: {e}")
            return "", f"error: {e}"

    def _parse_xls(self, file_path: Path) -> tuple[str, str]:
        import xlrd
        wb = xlrd.open_workbook(str(file_path))
        lines = [f"# {file_path.stem}\n"]
        for sheet in wb.sheets():
            lines.append(f"## {sheet.name}\n")
            lines.append(self._xls_sheet_to_md(sheet))
        return "\n".join(lines), "success"

    def _xls_sheet_to_md(self, sheet) -> str:
        rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        rows = [r for r in rows if any(v != "" for v in r)]
        return self._rows_to_md_tables(rows)

    def _parse_xlsx(self, file_path: Path) -> tuple[str, str]:
        import openpyxl
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        lines = [f"# {file_path.stem}\n"]
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"## {sheet_name}\n")
            rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
            rows = [r for r in rows if any(c is not None for c in r)]
            lines.append(self._rows_to_md_tables(rows))
        return "\n".join(lines), "success"

    def _rows_to_md_tables(self, rows) -> str:
        if not rows:
            return "（空表）\n"
        output: list[str] = []
        i = 0
        while i < len(rows):
            vals = self._row_vals(rows[i])
            if self._is_header(vals):
                output.append(f"### {vals[0].rstrip('：').rstrip(':')}\n")
                i += 1
                table_rows = []
                while i < len(rows):
                    rv = self._row_vals(rows[i])
                    if self._is_header(rv):
                        break
                    table_rows.append(rv)
                    i += 1
                if table_rows:
                    output.append(self._table_to_md(table_rows))
            else:
                table_rows = [vals]
                i += 1
                while i < len(rows):
                    rv = self._row_vals(rows[i])
                    if self._is_header(rv):
                        break
                    table_rows.append(rv)
                    i += 1
                output.append(self._table_to_md(table_rows))
        return "\n".join(output)

    @staticmethod
    def _row_vals(row) -> list[str]:
        return [str(v) if v is not None and v != "" else "" for v in row]

    @staticmethod
    def _is_header(vals: list[str]) -> bool:
        return bool(vals and vals[0] and not any(v for v in vals[1:]))

    def _table_to_md(self, rows: list[list[str]]) -> str:
        max_cols = max(len(r) for r in rows)
        lines = []
        lines.append("| " + " | ".join(rows[0][c] if c < len(rows[0]) else "" for c in range(max_cols)) + " |")
        lines.append("| " + " | ".join("---" for _ in range(max_cols)) + " |")
        for row in rows[1:]:
            cells = [row[c] if c < len(row) else "" for c in range(max_cols)]
            lines.append("| " + " | ".join(cells) + " |")
        lines.append("")
        return "\n".join(lines)


# ── DOCX → Markdown ───────────────────────────────────────────────────

class DocxParser:
    def parse(self, file_path: Path, image_dir: Path | None = None, doc_stem: str | None = None) -> tuple[str, str]:
        import docx
        from docx.oxml.ns import qn
        try:
            doc = docx.Document(str(file_path))
        except Exception as e:
            return "", f"error: {e}"

        body = doc.element.body
        lines: list[str] = [f"# {file_path.stem}\n"]
        _img_dir = image_dir
        _img_rel = f"images/{doc_stem}" if doc_stem else "images"
        if _img_dir:
            _img_dir.mkdir(parents=True, exist_ok=True)
        img_counter = 0

        for child in body:
            tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
            if tag == "p":
                para = docx.text.paragraph.Paragraph(child, doc)
                image_md = self._extract_images(child, doc, _img_dir, _img_rel, img_counter)
                if image_md:
                    img_counter += image_md.count("![")
                    lines.append(image_md)
                text = para.text.strip()
                if not text:
                    continue
                style = para.style.name if para.style else ""
                level = self._heading_level(style, child, qn)
                if level:
                    lines.append(f"{'#' * level} {text}\n")
                else:
                    runs_md = []
                    for run in para.runs:
                        t = run.text
                        if not t:
                            continue
                        if run.bold and run.italic:
                            t = f"***{t}***"
                        elif run.bold:
                            t = f"**{t}**"
                        elif run.italic:
                            t = f"*{t}*"
                        runs_md.append(t)
                    lines.append("".join(runs_md) if runs_md else text)
                    lines.append("")
            elif tag == "tbl":
                md = self._table_to_md(child, doc)
                if md:
                    lines.append(md)
                    lines.append("")
        return "\n".join(lines), "success"

    def _extract_images(self, child, doc, image_dir, rel_path, start_counter) -> str:
        from docx.oxml.ns import qn
        if not image_dir:
            return ""
        blips = child.findall(".//" + qn("a:blip"))
        if not blips:
            return ""
        lines = []
        for i, blip in enumerate(blips):
            rId = blip.get(qn("r:embed"))
            if not rId:
                continue
            try:
                image_part = doc.part.related_parts[rId]
                image_bytes = image_part.blob
                ext = image_part.content_type.split("/")[-1]
                if ext == "jpeg":
                    ext = "jpg"
            except Exception:
                continue
            idx = start_counter + i + 1
            fname = f"image{idx}.{ext}"
            out_path = image_dir / fname
            out_path.write_bytes(image_bytes)
            lines.append(f"![{fname}]({rel_path}/{fname})")
        return "\n".join(lines) if lines else ""

    def _heading_level(self, style_name: str, p_elem, qn) -> int:
        s = style_name.lower() if style_name else ""
        if s.startswith("heading") or s.startswith("标题"):
            for digit in s:
                if digit.isdigit():
                    return min(int(digit), 6)
            return 1
        pPr = p_elem.find(qn("w:pPr"))
        if pPr is not None:
            ol = pPr.find(qn("w:outlineLvl"))
            if ol is not None:
                return min(int(ol.get(qn("w:val"), "0")) + 1, 6)
        return 0

    def _table_to_md(self, tbl_elem, doc) -> str:
        import docx.table
        table = docx.table.Table(tbl_elem, doc)
        rows = [[cell.text.replace("\n", " ") for cell in row.cells] for row in table.rows]
        rows = [r for r in rows if any(c.strip() for c in r)]
        if not rows:
            return ""
        max_cols = max(len(r) for r in rows)
        lines = []
        lines.append("| " + " | ".join(rows[0] + [""] * (max_cols - len(rows[0]))) + " |")
        lines.append("| " + " | ".join("---" for _ in range(max_cols)) + " |")
        for row in rows[1:]:
            padded = row + [""] * (max_cols - len(row))
            lines.append("| " + " | ".join(padded) + " |")
        return "\n".join(lines)


# ── Kreuzberg parser ──────────────────────────────────────────────────

class KreuzbergParser:
    def parse(self, file_path: Path) -> tuple[str, str]:
        try:
            import asyncio
            from kreuzberg import extract_file
            result = asyncio.run(extract_file(str(file_path)))
            return f"# {file_path.stem}\n\n{result.content}", "success"
        except ImportError:
            return "", "error: kreuzberg not installed"
        except Exception as e:
            logger.warning(f"Kreuzberg parse failed for {file_path.name}: {e}")
            return "", f"error: {e}"


# ── Orchestrator ──────────────────────────────────────────────────────

class DocumentParser:
    DOCX_EXTS = {".docx"}
    PDF_PPT_EXTS = {".pdf", ".pptx", ".ppt", ".doc"}
    XLS_EXTS = {".xls", ".xlsx", ".xlsm"}
    TXT_EXTS = {".txt", ".md"}

    def __init__(self, use_kreuzberg: bool = True):
        self.use_kreuzberg = use_kreuzberg
        self.kreuzberg = KreuzbergParser() if use_kreuzberg else None
        self.zhipu = ZhipuFileParser()
        self.docx = DocxParser()
        self.xls = XlsParser()
        self._image_dir: Path | None = None
        self._doc_stem: str | None = None

    def set_image_context(self, image_dir: Path | None, doc_stem: str | None = None):
        self._image_dir = image_dir
        self._doc_stem = doc_stem

    def parse(self, file_path: Path) -> dict:
        ext = file_path.suffix.lower()
        result = {
            "filename": file_path.name, "filepath": str(file_path),
            "ext": ext, "parser": "", "status": "", "content": "", "char_count": 0,
        }

        if ext in self.DOCX_EXTS:
            result["parser"] = "python-docx"
            content, status = self.docx.parse(file_path, self._image_dir, self._doc_stem)
        elif ext in self.PDF_PPT_EXTS:
            if self.use_kreuzberg and self.kreuzberg:
                content, status = self.kreuzberg.parse(file_path)
                if status == "success" and content.strip():
                    result["parser"] = "kreuzberg"
                else:
                    result["parser"] = "zhipu"
                    content, status = self.zhipu.parse(file_path)
            else:
                result["parser"] = "zhipu"
                content, status = self.zhipu.parse(file_path)
        elif ext in self.XLS_EXTS:
            result["parser"] = "openpyxl_xlrd"
            content, status = self.xls.parse(file_path)
        elif ext in self.TXT_EXTS:
            result["parser"] = "txt_direct"
            content = file_path.read_text(encoding="utf-8", errors="replace")
            content = f"# {file_path.stem}\n\n{content}"
            result["content"] = content
            result["status"] = "success"
            result["char_count"] = len(content)
            return result
        else:
            result["parser"] = "unknown"
            result["status"] = f"error: unsupported format {ext}"
            return result

        result["content"] = content
        result["status"] = status
        result["char_count"] = len(content)
        return result
